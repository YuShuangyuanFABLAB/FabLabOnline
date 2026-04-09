# -*- coding: utf-8 -*-
"""
Excel导入模块 (F020, F021)
支持从Excel文件读取课程数据
"""

import os
from pathlib import Path
from typing import List, Dict, Any, Optional
from dataclasses import dataclass

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from src.core.models import (
    CourseUnitData, LayoutConfig,
    EvaluationLevel, OverallEvaluation, HomeworkEvaluation
)


@dataclass
class ExcelImportResult:
    """Excel导入结果"""
    success: bool
    data: List[CourseUnitData]
    errors: List[str]
    row_count: int
    valid_count: int


class ExcelTemplateGenerator:
    """Excel模板生成器"""

    # 列定义
    COLUMNS = [
        ("课时编号", "lesson_number", int),
        ("课程内容", "course_content", str),
        ("学生姓名", "student_name", str),
        ("授课教师", "teacher_name", str),
        ("课时数", "class_hours", int),
        ("上课日期", "class_date", str),
        ("逻辑思维", "logic_thinking", str),
        ("内容理解", "content_understanding", str),
        ("任务完成", "task_completion", str),
        ("倾听习惯", "listening_habit", str),
        ("问题解决", "problem_solving", str),
        ("独立分析", "independent_analysis", str),
        ("知识熟练度", "knowledge_proficiency", str),
        ("想象创造力", "imagination_creativity", str),
        ("抗挫折能力", "frustration_handling", str),
        ("学习方法", "learning_method", str),
        ("动手能力", "hands_on_ability", str),
        ("专注效率", "focus_efficiency", str),
        ("总体评价", "overall_evaluation", str),
        ("上次作业", "last_homework_status", str),
        ("补充说明", "additional_comments", str),
        ("课堂作业", "homework", str),
        ("注意事项", "other_notes", str),
    ]

    @classmethod
    def generate_template(cls, output_path: str) -> bool:
        """生成Excel模板文件"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "课程数据"

            # 设置样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 写入表头
            for col, (header, _, _) in enumerate(cls.COLUMNS, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # 写入示例数据
            example_data = [
                1, "学习机械臂基本结构", "张三", "李老师", 2, "2025年1月15日 14:00-16:00",
                "优", "良", "优", "良", "中", "良", "优", "良", "未体现", "良", "优", "良",
                "优", "良", "今天学习状态很好", "完成练习题", "下次带笔记本"
            ]

            for col, value in enumerate(example_data, 1):
                cell = ws.cell(row=2, column=col, value=value)
                cell.border = thin_border

            # 添加说明sheet
            ws2 = wb.create_sheet("填写说明")
            instructions = [
                ["课程反馈PPT导入模板说明"],
                [""],
                ["列名说明:"],
                ["- 课时编号: 数字，表示第几课"],
                ["- 课程内容: 课程的主要内容描述"],
                ["- 学生姓名: 学生姓名"],
                ["- 授课教师: 授课教师姓名"],
                ["- 课时数: 课程时长（课时）"],
                ["- 上课日期: 上课日期时间"],
                [""],
                ["评价等级（12项评价）:"],
                ["- 优/良/中/差/未体现"],
                [""],
                ["总体评价:"],
                ["- 优/良/仍需努力/需要改进"],
                [""],
                ["上次作业:"],
                ["- 优/良/中/差/无"],
                [""],
                ["注意事项:"],
                ["1. 第一行为表头，请勿修改"],
                ["2. 每一行代表一个课程单元"],
                ["3. 评价等级必须使用规定的选项"],
                ["4. 文件格式: .xlsx"],
            ]

            for row, row_data in enumerate(instructions, 1):
                ws2.cell(row=row, column=1, value=row_data[0] if row_data else "")

            # 调整列宽
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['F'].width = 25
            ws.column_dimensions['U'].width = 30
            ws.column_dimensions['V'].width = 20
            ws.column_dimensions['W'].width = 20

            wb.save(output_path)
            return True

        except Exception as e:
            print(f"生成模板失败: {e}")
            return False


class ExcelImporter:
    """Excel导入器"""

    # 评价等级映射
    LEVEL_MAP = {
        "优": EvaluationLevel.EXCELLENT,
        "良": EvaluationLevel.GOOD,
        "中": EvaluationLevel.MEDIUM,
        "差": EvaluationLevel.POOR,
        "未体现": EvaluationLevel.NOT_SHOWN,
    }

    OVERALL_MAP = {
        "优": OverallEvaluation.EXCELLENT,
        "良": OverallEvaluation.GOOD,
        "仍需努力": OverallEvaluation.NEED_EFFORT,
        "需要改进": OverallEvaluation.NEED_IMPROVEMENT,
    }

    HOMEWORK_MAP = {
        "优": HomeworkEvaluation.EXCELLENT,
        "良": HomeworkEvaluation.GOOD,
        "中": HomeworkEvaluation.MEDIUM,
        "差": HomeworkEvaluation.POOR,
        "无": HomeworkEvaluation.NONE,
    }

    def __init__(self):
        self.errors: List[str] = []

    def import_from_file(self, file_path: str) -> ExcelImportResult:
        """从Excel文件导入数据"""
        self.errors = []
        data_list = []

        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active

            # 获取列映射
            header_map = {}
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    header_map[header] = col

            # 读取数据行
            row_count = 0
            for row in range(2, ws.max_row + 1):
                row_count += 1
                try:
                    data = self._parse_row(ws, row, header_map)
                    if data:
                        data_list.append(data)
                except Exception as e:
                    self.errors.append(f"第{row}行解析错误: {e}")

            wb.close()

            return ExcelImportResult(
                success=True,
                data=data_list,
                errors=self.errors,
                row_count=row_count,
                valid_count=len(data_list)
            )

        except Exception as e:
            return ExcelImportResult(
                success=False,
                data=[],
                errors=[f"文件读取失败: {e}"],
                row_count=0,
                valid_count=0
            )

    def _parse_row(self, ws, row: int, header_map: Dict[str, int]) -> Optional[CourseUnitData]:
        """解析一行数据"""
        def get_value(key: str, default=None):
            col = header_map.get(key)
            if col:
                return ws.cell(row=row, column=col).value
            return default

        # 检查必填字段
        student_name = get_value("学生姓名")
        if not student_name:
            self.errors.append(f"第{row}行: 学生姓名为空")
            return None

        course_content = get_value("课程内容")
        if not course_content:
            self.errors.append(f"第{row}行: 课程内容为空")
            return None

        # 创建数据对象
        data = CourseUnitData()

        # 基本信息
        data.lesson_number = int(get_value("课时编号") or 1)
        data.course_content = str(course_content)
        data.student_name = str(student_name)
        data.teacher_name = str(get_value("授课教师") or "")
        data.class_hours = int(get_value("课时数") or 2)
        data.class_date = str(get_value("上课日期") or "")

        # 12项评价
        eval_fields = [
            "logic_thinking", "content_understanding", "task_completion",
            "listening_habit", "problem_solving", "independent_analysis",
            "knowledge_proficiency", "imagination_creativity", "frustration_handling",
            "learning_method", "hands_on_ability", "focus_efficiency"
        ]

        eval_names = [
            "逻辑思维", "内容理解", "任务完成", "倾听习惯", "问题解决", "独立分析",
            "知识熟练度", "想象创造力", "抗挫折能力", "学习方法", "动手能力", "专注效率"
        ]

        for field, name in zip(eval_fields, eval_names):
            value = get_value(name)
            if value and str(value) in self.LEVEL_MAP:
                setattr(data, field, self.LEVEL_MAP[str(value)])

        # 总体评价
        overall = get_value("总体评价")
        if overall and str(overall) in self.OVERALL_MAP:
            data.overall_evaluation = self.OVERALL_MAP[str(overall)]

        # 上次作业
        homework = get_value("上次作业")
        if homework and str(homework) in self.HOMEWORK_MAP:
            data.last_homework_status = self.HOMEWORK_MAP[str(homework)]

        # 文本字段
        data.additional_comments = str(get_value("补充说明") or "")
        data.homework = str(get_value("课堂作业") or "")
        data.other_notes = str(get_value("注意事项") or "")

        return data

    @staticmethod
    def validate_file(file_path: str) -> tuple:
        """验证文件格式"""
        if not os.path.exists(file_path):
            return False, "文件不存在"

        ext = Path(file_path).suffix.lower()
        if ext not in ['.xlsx', '.xls']:
            return False, "文件格式不正确，请使用.xlsx或.xls文件"

        try:
            wb = load_workbook(file_path, data_only=True)
            wb.close()
            return True, "文件格式正确"
        except Exception as e:
            return False, f"无法读取文件: {e}"
